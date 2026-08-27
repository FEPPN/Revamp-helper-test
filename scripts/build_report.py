#!/usr/bin/env python3
"""Assembles the multi-tab Excel report (SERP, Ahrefs, GSC, Competitors, [AI
Generative]) for the France revamp checklist process (papernest.com). Output
is entirely in French, matching the checklist and stakeholder audience.

Usage:
  python build_report.py \
    --keyword "octopus energy avis" \
    --brand "Octopus Energy" \
    --matching-json ahrefs_matching.json \
    --serp-json serp.json \
    --competitors-json competitors.json \
    --gsc-csv page_queries.csv \
    --page-url "https://www.papernest.com/demarches-energie/octopus-energy/avis/" \
    --ia-generative-json ia_generative.json \
    --out report.xlsx

All input files are JSON/CSV already produced by live Ahrefs MCP / SerpAPI /
Serper / gsc-connector calls (see SKILL.md for the full workflow) — this
script only assembles them, it never hand-retypes data.
"""

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="5A52FF", end_color="5A52FF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
TITLE_FONT = Font(bold=True, size=13)


def style_header(ws, row=1, ncols=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def intent_label(intents):
    if not intents:
        return "n/d"
    order = ["informational", "navigational", "commercial", "transactional", "branded", "local"]
    labels = {
        "informational": "Informationnelle",
        "navigational": "Navigationnelle",
        "commercial": "Commerciale",
        "transactional": "Transactionnelle",
        "branded": "Branded",
        "local": "Locale",
    }
    return " + ".join(labels[k] for k in order if intents.get(k))


def detailed_intent_fr(keyword, brand):
    """Classify a secondary keyword into a detailed French search-intent
    description. `brand` (e.g. "Octopus Energy") is injected into the
    generic/default cluster text and used to detect brand-specific vs.
    off-topic terms — never hardcode a brand name here."""
    kw = keyword.lower()
    brand_lower = brand.lower()

    account_terms = ["mon compte", "espace client", "téléphone", "telephone", "service client",
                      "numero", "numéro", "contact"]
    if any(t in kw for t in account_terms) and "avis" not in kw:
        return (f"COMPTE / ASSISTANCE — L'utilisateur est déjà client {brand} (ou en passe de le "
                "devenir) et cherche une action concrète (connexion espace client, numéro "
                "d'assistance, contact), pas une évaluation. Intention navigationnelle/"
                "transactionnelle, hors périmètre de la page \"avis\" : indique surtout un besoin de "
                "maillage interne vers la page contact/assistance.")

    generic_provider_terms = ["fournisseur d'électricité", "fournisseur électricité", "comparateur",
                               "moins cher", "meilleur fournisseur", "kwh", "tarif réglementé",
                               "offre électricité", "offre d'électricité", "offre gaz",
                               "que choisir électricité", "électricité verte", "energie verte",
                               "gaz et électricité"]
    if any(t in kw for t in generic_provider_terms) and brand_lower not in kw and "avis" not in kw:
        return (f"SHOPPING GÉNÉRIQUE — Requête de catégorie (non brandée) : l'utilisateur compare les "
                f"fournisseurs d'énergie en général, pas {brand} spécifiquement. Volume élevé mais "
                f"intention trop large pour être résolue par la seule page \"avis {brand}\" : utile "
                "comme bassin de trafic top-of-funnel pour du maillage vers des pages de comparaison, "
                "pas comme H2/H3 direct.")

    generic_trust_terms = ["trustpilot", "trust pilot", "truspilot", "trustpilote", "trust pilote",
                            "avis vérifiés", "avis verifies", "site fiable", "fiabilité site",
                            "verif site", "avis site", "avis entreprise", "review", "reviews"]
    if any(t in kw for t in generic_trust_terms) and brand_lower not in kw:
        return ("VALIDATION TIERCE PARTIE (générique) — L'utilisateur ne cherche pas la marque en "
                "particulier mais veut comprendre/consulter un agrégateur d'avis (Trustpilot ou "
                "similaire) pour vérifier la fiabilité d'une entreprise avant de lui faire confiance. "
                "Signal pertinent : la page doit citer explicitement Trustpilot/UFC-Que Choisir avec "
                "badge/lien pour capter ce besoin de tiers reconnus.")

    negative_terms = ["négatif", "negatif", "arnaque", "faux avis", "litige", "réclamation", "problème"]
    if any(t in kw for t in negative_terms):
        return ("DUE DILIGENCE / RECHERCHE DE SIGNAUX NÉGATIFS — L'utilisateur a déjà une opinion "
                "plutôt positive ou est indécis, mais cherche ACTIVEMENT les expériences négatives "
                "avant de s'engager : il veut savoir \"quel est le pire qui puisse arriver\" avec ce "
                "fournisseur (facture erronée, assistance lente, résiliation compliquée). Si la page "
                "n'aborde pas honnêtement les points faibles, l'utilisateur perd confiance et va lire "
                "Reddit/les forums, où le ton est plus critique.")

    ufc_60m_terms = ["ufc", "60 millions"]
    if any(t in kw for t in ufc_60m_terms):
        return ("VALIDATION INSTITUTIONNELLE — L'utilisateur cherche spécifiquement l'avis "
                "d'associations de consommateurs indépendantes (UFC-Que Choisir, 60 Millions de "
                "Consommateurs), perçues comme plus impartiales que Trustpilot (où l'on suspecte des "
                "avis achetés/gérés par la marque). Requête à forte intention de vérification "
                "pré-achat.")

    # Default cluster: branded "avis <brand>" intent
    return (f"PREUVE SOCIALE AVANT SOUSCRIPTION — L'utilisateur est déjà intéressé par {brand} "
            "(a vu une offre, un comparatif de prix, ou en a entendu parler) mais avant de souscrire "
            "un contrat, il veut la confirmation que d'autres personnes réelles ont eu une expérience "
            "positive avec ce fournisseur : il cherche des anecdotes concrètes sur le service client, "
            "la gestion des factures/régularisations et la fiabilité générale, pas simplement une note "
            "en étoiles. C'est une intention commerciale \"douce\", pilotée par la réduction du risque "
            "perçu plus que par le prix.")


def aide_blocks():
    """Contenu de l'onglet Aide sous forme de blocs neutres (pas de mise en
    forme) — un seul endroit où éditer le texte, consommé à la fois par le
    writer Excel (openpyxl, ci-dessous) et le writer Google Sheets."""
    return [
        {"type": "heading", "text": "🆘 Comment utiliser ce rapport"},

        {"type": "title", "text": "Ça sert à quoi, ce fichier ?"},
        {"type": "para", "height": 75, "text":
         "Tu vas revamper une page \"avis [marque]\" (ou une page démarche énergie). Ce fichier rassemble en "
         "un seul endroit tout ce qu'il faudrait normalement chercher à la main sur 4-5 outils différents : "
         "ce que Google affiche pour ce mot-clé, les vraies variantes que les gens tapent (avec leur volume), "
         "les stats réelles de la page sur Search Console, et ce que font les concurrents. Le but : que tu "
         "saches quoi écrire et pourquoi, pas juste \"écrire un article sur EDF\"."},
        {"type": "spacer"},

        {"type": "title", "text": "Les 5 onglets, en clair"},
        {"type": "bullet", "label": "SERP", "text":
         "ce que Google montre en tapant le mot-clé aujourd'hui : qui est en top 10, quelles "
         "questions les gens posent en plus (\"Autres questions posées\"), et si Google affiche déjà un "
         "résumé IA (AI Overview) — utile pour voir si ta page a une chance d'être citée ou si elle doit "
         "juste viser le classement classique."},
        {"type": "bullet", "label": "Ahrefs", "text":
         "toutes les variantes du mot-clé principal que les gens tapent réellement, avec le "
         "volume de recherche de chacune. Exemple : pour \"avis edf\", ça remonte aussi \"izi by edf avis\", "
         "\"offre tempo edf avis\", \"sowee edf avis\"... — des offres EDF que la page doit peut-être couvrir "
         "ou au moins lier, sinon elle rate ce trafic-là."},
        {"type": "bullet", "label": "GSC", "text":
         "les requêtes qui, AUJOURD'HUI, amènent déjà des clics sur la page depuis Google (Search "
         "Console). Ça te dit ce qui marche déjà — à ne surtout pas casser en revampant — et parfois des "
         "requêtes surprenantes que la page capte sans même les viser exprès."},
        {"type": "bullet", "label": "Concurrents", "text":
         "la structure (titres H1/H2/H3) des pages équivalentes chez Selectra, Kelwatt, "
         "Hellowatt et Fournisseurs-électricité. Sert à repérer ce qu'ils couvrent que ta page ne couvre "
         "pas encore, ou à l'inverse ce qu'ils font mal / n'ont pas — un angle à exploiter."},
        {"type": "bullet", "label": "Entités", "text":
         "les vrais noms (marques, régulateurs, associations, offres concurrentes...) que les "
         "concurrents citent et que ta page ne cite jamais. Différent d'Ahrefs : ça révèle des manques "
         "qu'AUCUN mot-clé ne signale. Détail complet et exemple juste en dessous — c'est l'onglet le "
         "moins intuitif des cinq, prends 2 minutes pour le lire."},
        {"type": "spacer"},

        {"type": "title", "text": "Les entités — pourquoi c'est aussi important que les mots-clés"},
        {"type": "para", "height": 75, "text":
         "Un mot-clé, c'est une PHRASE que quelqu'un tape. Une entité, c'est une VRAIE CHOSE que Google "
         "reconnaît dans un texte : une entreprise (EDF, Selectra), une institution (la Commission de "
         "régulation de l'énergie), une association (UFC-Que Choisir), un concurrent cité en comparaison "
         "(Engie, TotalEnergies). Google ne se contente pas de compter les mots-clés d'une page : il regarde "
         "aussi QUELLES VRAIES CHOSES elle mentionne, pour juger si elle traite le sujet en profondeur ou en "
         "surface."},
        {"type": "para", "height": 60, "text":
         "Le problème que ça résout : un mot-clé manquant, tu le VOIS (0 volume, 0 clic dans GSC). Une "
         "entité manquante, elle, ne se voit NULLE PART dans les autres onglets — aucun outil de mots-clés "
         "ne te dira \"ta page ne cite jamais la Commission de régulation de l'énergie alors que 3 de tes "
         "concurrents le font\". C'est exactement ce que fait l'onglet Entités."},
        {"type": "para", "height": 60, "text":
         "Exemple réel (test sur une page \"avis EDF\" vs Selectra) : Selectra mentionne \"Enercoop\", "
         "\"CLEEE\", \"FNCCR\", \"Gaz de Bordeaux\" — des acteurs réels du secteur énergie. Aucun de ces noms "
         "n'apparaît comme mot-clé à volume dans Ahrefs (donc invisible autrement), mais leur absence peut "
         "donner à Google l'impression que la page papernest couvre le sujet moins en profondeur que celle "
         "de Selectra."},
        {"type": "bullet", "label": "Colonne \"Présente chez\"", "text":
         "sur combien de concurrents (sur le total analysé) cette entité "
         "apparaît. Plus c'est haut, plus c'est un standard du sujet, pas un détail isolé chez un seul "
         "concurrent."},
        {"type": "bullet", "label": "Colonne \"Reconnue par Google\"", "text":
         "\"Oui\" = Google a formellement identifié cette entité comme "
         "une vraie organisation/marque connue (fiche Knowledge Graph). \"Vue mais pas identifiée\" = un "
         "nom propre repéré dans le texte mais que Google ne relie à aucune fiche connue — souvent un "
         "signal de bruit (nom de personne, marque trop petite) plutôt qu'une vraie priorité."},
        {"type": "bullet", "label": "3 sections du fichier", "text":
         "\"Manquantes chez nous\" (à considérer pour la réécriture, en priorité "
         "celles reconnues et présentes chez plusieurs concurrents) ; \"Partagées\" (le socle que tout le "
         "monde couvre déjà, rien à ajouter) ; \"Uniquement chez nous\" (soit un vrai différenciant à "
         "valoriser, soit hors-sujet à vérifier)."},
        {"type": "spacer"},

        {"type": "title", "text": "Les colonnes de l'onglet Ahrefs, une par une"},
        {"type": "bullet", "label": "Volume/mois", "text":
         "combien de personnes tapent EXACTEMENT cette phrase chaque mois. Simple : plus "
         "c'est haut, plus de gens cherchent ça mot pour mot."},
        {"type": "bullet", "label": "KD (Keyword Difficulty)", "text":
         "une note de 0 à 100 qui dit si c'est dur de se classer en top 10 "
         "Google sur ce mot précis (calculée sur la force des liens des pages déjà en tête, pas sur la "
         "qualité du contenu). 0-10 = facile à viser, 40+ = déjà très disputé, difficile sans gros travail."},
        {"type": "bullet", "label": "CPC ($)", "text":
         "le prix moyen qu'un annonceur paierait pour une pub sur ce mot-clé. Sert d'indice : "
         "plus c'est cher, plus la personne qui tape ça est proche d'acheter/souscrire."},
        {"type": "bullet", "label": "Potentiel de trafic", "text":
         "PAS le volume du mot-clé lui-même : c'est le trafic TOTAL (toutes requêtes "
         "confondues) que récolte la page actuellement n°1 dessus. Exemple : \"assistance dépannage edf "
         "avis\" a 150 recherches/mois, mais un Potentiel de trafic de 3100 — la page n°1 vit aussi de "
         "dizaines de variantes proches. Un potentiel très supérieur au volume = signal qu'il faut une "
         "vraie section dédiée à ce sujet, pas juste une phrase en passant."},
        {"type": "bullet", "label": "Intention (Ahrefs)", "text":
         "la catégorie officielle Ahrefs (Informationnelle, Commerciale, Branded...). "
         "Basique, utile pour trier vite."},
        {"type": "bullet", "label": "Intention de recherche (description détaillée)", "text":
         "généré spécifiquement pour ce rapport : "
         "explique EN FRANÇAIS SIMPLE pourquoi la personne cherche ce mot-clé et donc quoi écrire pour "
         "répondre à ce vrai besoin (ex: \"cherche des signaux négatifs avant de s'engager\" → il faut "
         "aborder honnêtement 1-2 points faibles, pas juste vanter la marque)."},
        {"type": "bullet", "label": "Parent Topic", "text":
         "le sujet plus large auquel Ahrefs rattache ce mot-clé. Utile pour regrouper "
         "plusieurs lignes qui parlent en fait de la même chose."},
        {"type": "spacer"},

        {"type": "title", "text": "Les colonnes de l'onglet GSC"},
        {"type": "bullet", "label": "Clics / Impressions", "text":
         "clics = personnes qui ont cliqué sur la page depuis Google ; impressions "
         "= nombre de fois où la page est juste apparue dans les résultats (cliquée ou non)."},
        {"type": "bullet", "label": "CTR", "text":
         "le % de clics par rapport aux impressions. Un CTR très bas avec beaucoup d'impressions "
         "= la page apparaît souvent mais le titre/résumé Google n'attire pas assez le clic."},
        {"type": "bullet", "label": "Position moyenne", "text":
         "le classement moyen de la page sur cette requête (1 = tout en haut)."},
        {"type": "spacer"},

        {"type": "title", "text": "Comment s'en servir, étape par étape"},
        {"type": "bullet", "label": "1.", "text":
         "Ouvre l'onglet Concurrents en premier — repère un angle ou un sujet qu'eux couvrent et "
         "pas la page papernest actuellement."},
        {"type": "bullet", "label": "2.", "text":
         "Ouvre GSC — note les 3-5 requêtes qui ramènent déjà le plus de clics : ne les perds "
         "jamais en réécrivant la page."},
        {"type": "bullet", "label": "3.", "text":
         "Ouvre Ahrefs, trie mentalement par Volume/mois décroissant (déjà fait dans le fichier) — "
         "repère les sous-thèmes/sous-marques avec du volume propre et un Potentiel de trafic élevé : "
         "ce sont tes candidats pour de nouvelles sections H2."},
        {"type": "bullet", "label": "4.", "text":
         "Pour chaque sous-thème retenu, lis sa colonne \"Intention de recherche\" — ça te dit quel "
         "angle rédiger (preuve sociale, réponse à une inquiétude, comparaison...)."},
        {"type": "bullet", "label": "5.", "height": 60, "text":
         "Ouvre l'onglet Entités — ajoute dans ta rédaction les entités \"Manquantes chez nous\" "
         "reconnues par Google et présentes chez plusieurs concurrents (régulateurs, associations, offres "
         "concurrentes citées en comparaison...) : invisibles dans Ahrefs, mais Google les compte quand "
         "même comme faisant partie du sujet complet."},
        {"type": "bullet", "label": "6.", "height": 60, "text":
         "Rédige/révise la page avec ces sections, en gardant les mots-clés GSC qui marchent déjà "
         "et en couvrant les angles concurrents repérés à l'étape 1."},
    ]


def build_aide_sheet(wb):
    ws = wb.active
    ws.title = "Aide"

    r = 1
    for block in aide_blocks():
        t = block["type"]
        if t == "heading":
            ws.cell(row=r, column=1, value=block["text"]).font = Font(bold=True, size=16, color="5A52FF")
            r += 2
        elif t == "title":
            ws.cell(row=r, column=1, value=block["text"]).font = Font(bold=True, size=13, color="5A52FF")
            r += 1
        elif t == "para":
            cell = ws.cell(row=r, column=1, value=block["text"])
            cell.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.row_dimensions[r].height = block.get("height", 45)
            r += 1
        elif t == "bullet":
            cell = ws.cell(row=r, column=1, value=f"{block['label']} — {block['text']}")
            cell.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.row_dimensions[r].height = block.get("height", 45)
            r += 1
        elif t == "spacer":
            r += 1

    autosize(ws, [110, 20, 20, 20])
    ws.sheet_view.showGridLines = False


def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def build_serp_sheet(wb, serp_path):
    data = load_json(serp_path)
    ws = wb.create_sheet("SERP")

    ws["A1"] = f"SERP — \"{data['keyword']}\" ({data.get('market', 'FR (google.fr)')})"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    headers = ["Position", "Titre", "URL", "Résumé"]
    hrow = 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, row=hrow, ncols=4)

    r = hrow + 1
    for item in data["organic"]:
        ws.cell(row=r, column=1, value=item["position"])
        ws.cell(row=r, column=2, value=item["title"]).alignment = WRAP
        ws.cell(row=r, column=3, value=item["link"]).alignment = WRAP
        ws.cell(row=r, column=4, value=item["note"]).alignment = WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Autres questions posées (PAA)").font = Font(bold=True)
    r += 1
    for q in data.get("paa", []):
        ws.cell(row=r, column=1, value=f"• {q}")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Recherches associées").font = Font(bold=True)
    r += 1
    for q in data.get("related_searches", []):
        ws.cell(row=r, column=1, value=f"• {q}")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Aperçu IA (AI Overview)").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Présent" if data.get("ai_overview_present") else "Absent")
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value=data.get("ai_overview_summary", "")).alignment = WRAP
    ws.row_dimensions[r].height = 90

    r += 2
    ws.cell(row=r, column=1, value="Knowledge Graph").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Présent" if data.get("knowledge_graph_present") else "Absent")

    autosize(ws, [12, 45, 55, 70])
    ws.freeze_panes = "A4"


def build_ahrefs_sheet(wb, matching_path, brand):
    # Ahrefs Keywords Explorer -> [seed kw] -> Keyword ideas -> Matching terms -> Terms match -> All terms
    rows = load_json(matching_path)
    rows = sorted(rows, key=lambda r: (r.get("volume") or 0), reverse=True)

    ws = wb.create_sheet("Ahrefs")
    ws["A1"] = "Ahrefs Keywords Explorer — Matching terms (Terms match, All terms)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    headers = ["Mot-clé", "Volume/mois", "KD", "CPC ($)", "Potentiel de trafic",
               "Intention (Ahrefs)", "Intention de recherche (description détaillée)",
               "Parent Topic"]
    hrow = 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, row=hrow, ncols=len(headers))

    for r, row in enumerate(rows, start=hrow + 1):
        ws.cell(row=r, column=1, value=row["keyword"])
        ws.cell(row=r, column=2, value=row.get("volume"))
        ws.cell(row=r, column=3, value=row.get("difficulty"))
        ws.cell(row=r, column=4, value=row.get("cpc"))
        ws.cell(row=r, column=5, value=row.get("traffic_potential"))
        ws.cell(row=r, column=6, value=intent_label(row.get("intents")))
        cell = ws.cell(row=r, column=7, value=detailed_intent_fr(row["keyword"], brand))
        cell.alignment = WRAP
        ws.cell(row=r, column=8, value=row.get("parent_topic"))
        ws.row_dimensions[r].height = 60

    autosize(ws, [40, 12, 8, 10, 16, 22, 90, 28])
    ws.freeze_panes = "A4"


def build_gsc_sheet(wb, gsc_csv_path, page_url):
    with open(gsc_csv_path, encoding="utf-8") as f:
        reader = list(csv.DictReader(f))

    rows = [r for r in reader if float(r["clicks"]) > 0]
    rows.sort(key=lambda r: float(r["clicks"]), reverse=True)

    ws = wb.create_sheet("GSC")
    ws["A1"] = f"GSC — {page_url}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    headers = ["Requête", "Clics", "Impressions", "CTR", "Position moyenne"]
    hrow = 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, row=hrow, ncols=len(headers))

    for r, row in enumerate(rows, start=hrow + 1):
        ws.cell(row=r, column=1, value=row["query"])
        ws.cell(row=r, column=2, value=int(float(row["clicks"])))
        ws.cell(row=r, column=3, value=int(float(row["impressions"])))
        ws.cell(row=r, column=4, value=round(float(row["ctr"]) * 100, 2))
        ws.cell(row=r, column=5, value=round(float(row["position"]), 2))

    autosize(ws, [45, 10, 14, 10, 16])
    ws.freeze_panes = "A4"


def build_competitor_sheet(wb, competitors_path):
    data = load_json(competitors_path)
    ws = wb.create_sheet("Concurrents")

    r = 1
    for comp in data:
        ws.cell(row=r, column=1, value=comp["site"]).font = Font(bold=True, size=12, color="5A52FF")
        r += 1
        ws.cell(row=r, column=1, value="URL :")
        ws.cell(row=r, column=2, value=comp["url"])
        r += 1
        ws.cell(row=r, column=1, value="H1 :")
        cell = ws.cell(row=r, column=2, value=comp["h1"])
        cell.font = Font(bold=True)
        r += 1
        if comp.get("no_dedicated_avis_page"):
            ws.cell(row=r, column=1, value="⚠️ Remarque")
            note_cell = ws.cell(row=r, column=2, value=comp["summary"])
            note_cell.alignment = WRAP
            note_cell.font = Font(italic=True, color="FF2056")
            r += 2
        for block in comp["structure"]:
            ws.cell(row=r, column=2, value=f"H2 : {block['h2']}").font = Font(bold=True)
            r += 1
            for h3 in block["h3"]:
                ws.cell(row=r, column=3, value=f"H3 : {h3}").alignment = WRAP
                r += 1
        r += 1
        if not comp.get("no_dedicated_avis_page"):
            ws.cell(row=r, column=1, value="Résumé :")
            cell = ws.cell(row=r, column=2, value=comp["summary"])
            cell.alignment = WRAP
            r += 1
        r += 2

    autosize(ws, [14, 70, 60])


def build_entity_sheet(wb, gap_data, page_url):
    """Écart d'entités (Google NLP) : ce que les concurrents mentionnent comme
    marques/produits/organisations reconnus et que notre page ne mentionne pas
    du tout — un signal de contenu manquant que les mots-clés seuls ne voient
    pas toujours (une entité peut manquer sans qu'aucun mot-clé Ahrefs ne le
    signale explicitement)."""
    ws = wb.create_sheet("Entités")
    n = gap_data.get("competitors", 0)

    ws["A1"] = f"Écart d'entités — {page_url}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws["A2"] = f"Comparé à {n} page(s) concurrente(s) via Google Natural Language API"
    ws["A2"].font = Font(italic=True, color="6B7280")

    r = 4
    ws.cell(row=r, column=1, value="⚠️ Manquantes chez nous, présentes chez les concurrents").font = Font(
        bold=True, size=12, color="FF2056")
    r += 1
    headers = ["Entité", "Présente chez", "Reconnue par Google"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, row=r, ncols=3)
    r += 1
    if gap_data.get("gaps"):
        for g in gap_data["gaps"]:
            ws.cell(row=r, column=1, value=g["name"])
            ws.cell(row=r, column=2, value=f"{g['on_competitors']}/{n} concurrent(s)")
            ws.cell(row=r, column=3, value="Oui" if g.get("mid") else "Vue mais pas identifiée")
            r += 1
    else:
        ws.cell(row=r, column=1, value="(aucun écart détecté)")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="✅ Partagées (nous + concurrents) — le socle attendu").font = Font(
        bold=True, size=12, color="5A52FF")
    r += 1
    if gap_data.get("shared"):
        cell = ws.cell(row=r, column=1,
                        value=", ".join(f"{s['name']} ({s['on_competitors']}/{n})" for s in gap_data["shared"]))
        cell.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 45
    else:
        ws.cell(row=r, column=1, value="(aucune)")
    r += 2

    ws.cell(row=r, column=1, value="🟢 Uniquement chez nous — différenciant ou hors-sujet à vérifier").font = Font(
        bold=True, size=12, color="02C5AE")
    r += 1
    if gap_data.get("our_unique"):
        cell = ws.cell(row=r, column=1, value=", ".join(e["name"] for e in gap_data["our_unique"]))
        cell.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 45
    else:
        ws.cell(row=r, column=1, value="(aucune)")

    autosize(ws, [40, 22, 26])
    ws.freeze_panes = "A5"


def build_ia_generative_sheet(wb, ia_path):
    data = load_json(ia_path)
    ws = wb.create_sheet("IA générative")

    r = 1
    ws.cell(row=r, column=1, value="Statut AI Overview (Google)").font = Font(bold=True, size=12, color="5A52FF")
    r += 1
    cell = ws.cell(row=r, column=1, value=data["ai_overview_status"])
    cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 60
    r += 2

    block = data["simulated_answer_no_browsing"]
    ws.cell(row=r, column=1, value=block["title"]).font = Font(bold=True, size=12, color="5A52FF")
    r += 1
    cell = ws.cell(row=r, column=1, value=block["text"])
    cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 120
    r += 2

    block = data["real_cited_sources_with_browsing"]
    ws.cell(row=r, column=1, value=block["title"]).font = Font(bold=True, size=12, color="5A52FF")
    r += 1
    for s in block["sources"]:
        ws.cell(row=r, column=1, value=f"• {s}").alignment = WRAP
        r += 1
    cell = ws.cell(row=r, column=1, value=block["note"])
    cell.font = Font(italic=True)
    cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 45
    r += 2

    block = data["hypothetical_fan_out_queries"]
    ws.cell(row=r, column=1, value=block["title"]).font = Font(bold=True, size=12, color="5A52FF")
    r += 1
    for q in block["queries"]:
        ws.cell(row=r, column=1, value=f"• {q}")
        r += 1
    if block.get("logic"):
        r += 1
        cell = ws.cell(row=r, column=1, value=block["logic"])
        cell.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 90
        r += 1
    cell = ws.cell(row=r, column=1, value=f"⚠️ {block['caveat']}")
    cell.font = Font(italic=True, color="FF2056")
    cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 45

    autosize(ws, [90, 20, 20])


def main():
    parser = argparse.ArgumentParser(description="Assemble the FR SEO multi-tab report (xlsx)")
    parser.add_argument("--keyword", required=True, help="Main keyword (FR)")
    parser.add_argument("--brand", required=True,
                         help="Brand/provider name as it should appear in the intent descriptions "
                              "(e.g. 'Octopus Energy', 'Engie', 'EDF') — never hardcode a brand in the script")
    parser.add_argument("--matching-json", required=True, help="JSON from Ahrefs keywords-explorer-matching-terms")
    parser.add_argument("--serp-json", required=True, help="Structured SERP JSON (SerpAPI)")
    parser.add_argument("--competitors-json", required=True, help="H1-H3 competitor structure JSON")
    parser.add_argument("--gsc-csv", required=True, help="gsc-connector CSV filtered with --page")
    parser.add_argument("--page-url", required=True, help="Exact target papernest.com page URL")
    parser.add_argument("--ia-generative-json", help="Optional: AI Generative tab JSON (AI Overview + fan-out)")
    parser.add_argument("--entity-gap-json", help="Optional: entity gap JSON (analyze_entities.py --gap-vs --json)")
    parser.add_argument("--out", required=True, help="Output xlsx path")
    args = parser.parse_args()

    wb = Workbook()
    build_aide_sheet(wb)
    build_serp_sheet(wb, args.serp_json)
    build_ahrefs_sheet(wb, args.matching_json, args.brand)
    build_gsc_sheet(wb, args.gsc_csv, args.page_url)
    build_competitor_sheet(wb, args.competitors_json)
    if args.entity_gap_json:
        build_entity_sheet(wb, load_json(args.entity_gap_json).get("gap", {}), args.page_url)
    if args.ia_generative_json:
        build_ia_generative_sheet(wb, args.ia_generative_json)

    out_path = Path(args.out)
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
